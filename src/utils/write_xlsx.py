from openpyxl import Workbook  # new import
from openpyxl.styles import PatternFill  # NEW
from pathlib import Path 



def write_results_xlsx(records, RESULTS_PATH):
    # Save results to Excel in a 4x3 layout per plate
    # Excel styles
    TITLE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # sky blue
    LABEL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # grey

    wells_per_plate = 12  # 4 columns * 3 rows
    rows_per_plate = 3
    cols_per_plate = 4
    num_plates = len(records) // wells_per_plate  # assuming perfect 4x3 plates

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    current_row = 1
    column_headers = [4, 3, 2, 1]
    row_labels = ["A", "B", "C"]

    for plate_idx in range(num_plates):
        start = plate_idx * wells_per_plate
        plate_records = records[start:start + wells_per_plate]

        # Build plate title from first well filename:
        raw_name = plate_records[0].get("name", "unknown")
        plate_title = Path(raw_name).stem  # remove extension
        plate_title = plate_title.removesuffix("_A1")
        plate_title = plate_title.removeprefix("cropped_plates_")

        # Title row in sky blue across 5 columns
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            if col == 1:
                cell.value = plate_title
            cell.fill = TITLE_FILL
        current_row += 1

        # Header row: 4 3 2 1 (in grey)
        for col_idx, col_label in enumerate(column_headers, start=1):
            header_cell = ws.cell(row=current_row, column=col_idx, value=col_label)
            header_cell.fill = LABEL_FILL
        current_row += 1

        # 3 rows (A, B, C), 4 columns each, plus row label at the end
        for r in range(rows_per_plate):
            for c in range(cols_per_plate):
                idx = r * cols_per_plate + c
                count_value = plate_records[idx]["predicted_count"]
                if count_value == -1:
                    count_value = "uncountable"
                # inverse writing
                excel_col = cols_per_plate - c  # 0->4, 1->3, 2->2, 3->1
                ws.cell(row=current_row, column=excel_col, value=count_value)

            # Row label (A/B/C) in grey
            label_cell = ws.cell(row=current_row, column=cols_per_plate + 1, value=row_labels[r])
            label_cell.fill = LABEL_FILL
            current_row += 1

        # Blank line between plates
        current_row += 1

    wb.save(RESULTS_PATH)